from PyPDF2 import PdfFileWriter, PdfFileReader
from pathlib import Path
from openpyxl import load_workbook

quran_path = r"files\arabic-quran.pdf"
book_path = r"files\bookarabic.pdf"

#QURAN
wb = load_workbook("files\data.xlsx")
ws = wb.active

input_pdf = PdfFileReader(str(quran_path))
#add 2 pages for actual page

pdf_writer = PdfFileWriter()

quran_page = ws['C2'].value

for n in range(quran_page,quran_page + 2):
    page = input_pdf.getPage(n)
    pdf_writer.addPage(page)

day = ws['A2'].value
quran_day = str(day)
quran_name = "quran_day" + quran_day + ".pdf"

with Path(quran_name).open(mode="wb") as output_file:
    pdf_writer.write(output_file)

ws['C2'] = ws['C2'].value + 2

#BOOK

book_day = str(day)
book_name = "book_day" + book_day + ".pdf"

book_page = ws['B2'].value

input_pdf = PdfFileReader(str(book_path))
pdf_writer = PdfFileWriter()

for n in range(book_page,book_page + 2):
    page = input_pdf.getPage(n)
    pdf_writer.addPage(page)

with Path(book_name).open(mode="wb") as output_file:
    pdf_writer.write(output_file)

ws['B2'] = ws['B2'].value + 2

ws['A2'] = ws['A2'].value + 1
wb.save('files\data.xlsx')